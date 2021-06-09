import random
import set_functions
import subprocess
import time
import os
import SGAC
from openpyxl import Workbook
from openpyxl import load_workbook

for number in range(0, 5):
    Python_Sub_Graph = dict()
    Python_Res_Graph = dict()
    Python_Sub_Closure_Graph = dict()
    Python_Res_Closure_Graph = dict()
    Python_Rules = dict()
    Python_Context = []

    #minimum_subject_nodes = 5
    maximum_subject_nodes = 20
    maximum_resource_nodes = 20
    #minimum_resource_node = 5
    #minimum_rules = 13
    maximum_rules = 30
    #minimum_contexts = 10
    maximum_contexts = 10
    biggestCon = 0

    #for node in range(0, random.randrange(minimum_resource_node, maximum_subject_nodes, 1)):
    for node in range(0, maximum_subject_nodes):
        Python_Sub_Graph["s"+str(node)] = []
        for childNode in range(0, node):
            if random.randint(0, 1):
                Python_Sub_Graph["s"+str(node)].append("s"+str(childNode))

    for node in range(0, maximum_resource_nodes):
        Python_Res_Graph["r"+str(node)] = []
        for childNode in range(0, node):
            if random.randint(0, 1):
                Python_Res_Graph["r"+str(node)].append("r"+str(childNode))

    for context in range(0, maximum_contexts):
        Python_Context.append("c"+str(context))

    for rule in range(0, maximum_rules):
        Python_Rules["rule"+str(rule)] = []
        Python_Rules["rule"+str(rule)].append("s"+str(random.randint(0, len(Python_Sub_Graph.keys())-1)))  # Subject
        Python_Rules["rule"+str(rule)].append("r"+str(random.randint(0, len(Python_Res_Graph.keys())-1)))  # Resource
        if random.randint(0, 1):
            Python_Rules["rule" + str(rule)].append("prohibition")  # Modality
        else:
            Python_Rules["rule" + str(rule)].append("permission")  # Modality
        Python_Rules["rule"+str(rule)].append(random.randint(0, 4))  # Priority
        contextsVector = []
        for i in range(0, random.randint(1, len(Python_Context))):
            randomInt = random.randint(0, len(Python_Context)-1)
            randomContext = "c"+str(randomInt)
            if randomContext not in contextsVector:
                contextsVector.append(randomContext)
                if randomInt > biggestCon:
                    biggestCon = randomInt
        Python_Rules["rule"+str(rule)].append(contextsVector)

    set_functions.transitive_closure(Python_Sub_Graph, Python_Sub_Closure_Graph)
    set_functions.transitive_closure(Python_Res_Graph, Python_Res_Closure_Graph)

    tab = "    "
    V_SUB = ""
    V_RES = ""
    RULE_T = ""
    graph_res = ""
    graph_sub = ""
    rules = ""
    CONTEXT = ""
    MODALITY = "{per, pro}"
    rulesInAlloy = []

    for key in Python_Sub_Graph.keys():
        V_SUB += str(key)+", "
        for node in Python_Sub_Graph[key]:
            graph_sub += str(key)+"|->"+str(node)+", "

    for key in Python_Res_Graph.keys():
        V_RES += str(key)+", "
        for node in Python_Res_Graph[key]:
            graph_res += str(key)+"|->"+str(node)+", "

    for i in range(0, biggestCon+1):
        CONTEXT += "c"+str(i)+", "

    for key in Python_Rules.keys():
        eachAlloyRule = []
        eachAlloyRule.append("one sig "+key+" extends Rule{}{")
        RULE_T += key+", "
        rules += key+"|->(rec("
        rules += "su:"+Python_Rules[key][0]+","
        eachAlloyRule.append(tab + "s =" + Python_Rules[key][0])
        rules += "re:"+Python_Rules[key][1]+","
        eachAlloyRule.append(tab + "t =" + Python_Rules[key][1])
        if Python_Rules[key][2] == "permission":
            rules += "mo:per,"
            eachAlloyRule.append(tab + "m = permission")
        else:
            rules += "mo:pro,"
            eachAlloyRule.append(tab + "m = prohibition")
        rules += "pr: "+str(Python_Rules[key][3])+","
        eachAlloyRule.append(tab + "p = " + str(Python_Rules[key][3]))
        rules += "co:{"
        conList = ""
        for con in Python_Rules[key][4]:
            rules += con+","
            conList += con+"+"
        rules = rules[:len(rules)-1:]+"})), "
        eachAlloyRule.append(tab + "c = " + conList[:len(conList)-1:])
        rulesInAlloy.append(eachAlloyRule)

    bmachine = "/*test/SGAC.mch\nAuthor: Diego Oliveira\n*/\n\n"
    bmachine += "MACHINE\n"+tab+"SGAC\n\nDEFINITIONS\n"+tab+"SET_PREF_MAX_OPERATIONS == 1000;\n"
    bmachine += tab+"applicable_def(req) == {rul | is_applicable(req,rul)};\n"\
                +tab+"is_applicable(req,rul) == ( rul : RULE_T & dom({req}) <: cl_e_sub[{(rules(rul))'su}] \\/ " \
                     "{(rules(rul))'su} & ran({req}) <: cl_e_res[{(rules(rul))'re}] \\/ {(rules(rul))'re});\n"
    bmachine += tab+"maxElem(req) == {rul | rul : applicable(req) & not(#rul2.(rul2:applicable(req) & rul|->rul2: " \
                    "lessSpecific))};\n"
    bmachine += tab+"access_def(req,con) == (!rsinks.(rsinks: pseudoSink(req,con) => (rules(rsinks))'mo = per) & " \
                    "pseudoSink(req,con)/={});\n"
    bmachine += tab+"ruleSucc == %xx.(xx:REQUEST_T | {yy,zz| yy:applicable(xx) & zz:applicable(xx) & yy|->zz : " \
                    "isPrecededBy(xx) & not(#uu.(uu : RULE_T & yy |-> uu : isPrecededBy(xx) & uu |-> zz : " \
                    "isPrecededBy(xx) & uu /= yy & uu /= zz))})\n\n"
    bmachine += "SETS\n"+tab+"/*context type*/\n"+tab+"CONTEXT={"+CONTEXT[:len(CONTEXT)-2:]+"};\n"
    bmachine += tab+"/*subject vertex type*/\n"+tab+"V_SUB={"+V_SUB[:len(V_SUB)-2:]+"};\n"
    bmachine += tab+"/*resource vertex type*/\n"+tab+"V_RES={"+V_RES[:len(V_RES)-2:]+"};\n"
    bmachine += tab+"/*rule type*/\n"+tab+"RULE_T={"+RULE_T[:len(RULE_T)-2:]+"};\n"
    bmachine += tab+"/* modality type*/\n"+tab+"MODALITY="+MODALITY+"\n\n"
    bmachine += "CONSTANTS\n"+tab+"/*set of all requests*/\n"+tab+"REQUEST_T,\n"
    bmachine += tab+"/*set of rules of the policy*/\n"+tab+"rules,\n"
    bmachine += tab+"/*edges of the subject graph*/\n"+tab+"e_sub,\n"
    bmachine += tab+"/*resource graph edges*/\n"+tab+"e_res,\n"
    bmachine += tab+"/*ordering 1: lessSpecific*/\n"+tab+"lessSpecific,\n"
    bmachine += tab+"/*closure1 of e_sub, closure of e_sub, e_res*/\n"+tab+"cl1_e_sub,cl_e_sub,cl1_e_res,cl_e_res\n\n"
    bmachine += "PROPERTIES\n"+tab+"/*types*/\n"+tab+"e_sub: V_SUB <-> V_SUB &\n"
    bmachine += tab+"e_res : V_RES <-> V_RES &\n"
    bmachine += tab+"REQUEST_T = (V_SUB-dom(e_sub)) * (V_RES-dom(e_res)) &\n"
    bmachine += tab+"/*rules: function that maps a rule to the rule structure*/\n"
    bmachine += tab+"rules: RULE_T --> (struct(su:V_SUB, re:V_RES, mo: MODALITY, pr:INTEGER, co:POW(CONTEXT))) &\n"
    bmachine += tab+"lessSpecific : RULE_T <-> RULE_T & \n\n"
    bmachine += tab+"/* closure definitions */\n"
    bmachine += tab+"cl1_e_sub = closure1(e_sub) &\n"
    bmachine += tab+"cl_e_sub = closure(e_sub) &\n"
    bmachine += tab+"cl1_e_res = closure1(e_res) &\n"
    bmachine += tab+"cl_e_res = closure(e_res) &\n"
    bmachine += tab+"/*acyclicity of the graphs */\n"
    bmachine += tab+"cl1_e_sub /\\ id(V_SUB) = {} &\n"
    bmachine += tab+"cl1_e_res /\\ id(V_RES) = {} &\n\n"
    bmachine += tab+"/* Constraints for rule ordering */\n"
    bmachine += tab+"/* lessSpecific definition */\n"
    bmachine += tab+"lessSpecific = {xx,yy | xx: dom(rules) & yy: dom(rules) & ((((rules(xx))'pr > (rules(yy))'pr) or " \
                    "(((rules(xx))'pr = (rules(yy))'pr) & (rules(yy))'su: cl1_e_sub[{(rules(xx))'su}])))}&\n\n"
    bmachine += tab+"/* pseudo INITIALISATION  */\n"
    bmachine += tab+"/* THE GRAPHS */\n"
    bmachine += tab+"e_sub={"+graph_sub[:len(graph_sub)-2:]+"} &\n"
    bmachine += tab+"e_res={"+graph_res[:len(graph_res)-2:]+"} &\n"
    bmachine += tab+"rules={"+rules[:len(rules)-2:]+"}\n\n"
    bmachine += "VARIABLES\n"
    bmachine += tab+"/*applicable: contains all applicable rule to a request*/\n"+tab+"applicable,\n"
    bmachine += tab+"/*conRule: associate a condition to its rules*/\n"+tab+"conRule,\n"
    bmachine += tab+"/*ordering 2:isPrecededBy*/\n"+tab+"isPrecededBy,\n"
    bmachine += tab+"/*closure of ruleSucc*/\n"+tab+"cl1_ruleSucc,\n"
    bmachine += tab+"/*function returning the pseudosinks of a request+context*/\n"+tab+"pseudoSink\n\n"
    bmachine += "INVARIANT\n"+tab+"applicable :  REQUEST_T -->POW(RULE_T) &\n"
    bmachine += tab+"conRule : CONTEXT --> POW(RULE_T) &\n"+tab+"isPrecededBy : REQUEST_T --> (RULE_T <-> RULE_T) &\n"
    bmachine += tab+"cl1_ruleSucc : REQUEST_T --> (RULE_T <-> RULE_T) &\n"+tab+"pseudoSink : (REQUEST_T * CONTEXT)-->" \
                                                                               " POW(RULE_T)\n\n"
    bmachine += "INITIALISATION\n"+tab+"BEGIN\n"+tab+"applicable :=  %rr.(rr:REQUEST_T|applicable_def(rr));\n"
    bmachine += tab+"conRule := %con.(con:CONTEXT|{cc|cc:dom(rules) & con : (rules(cc))'co});\n"
    bmachine += tab+"/*isPrecededBy definition*/\n"
    bmachine += tab+"isPrecededBy := %xx.(xx:REQUEST_T | {yy, zz | yy:applicable(xx) & zz:applicable(xx) & yy/=zz & (" \
                    "yy|->zz : lessSpecific or ({yy,zz}<:maxElem(xx) & (rules(yy))'mo = per & (rules(zz))'mo = pro))});\n"
    bmachine += tab+"cl1_ruleSucc := %xx.(xx:REQUEST_T | closure1(isPrecededBy(xx)));\n"
    bmachine += tab+"pseudoSink := %(req,con).(req:REQUEST_T & con:dom(conRule) | {ru | ru : applicable(req) &" \
                    "ru : conRule(con) & !subrule.(subrule : cl1_ruleSucc(req)[{ru}] => not( subrule: conRule(con)))})\n"
    bmachine += tab+"END\n\n"
    bmachine += "OPERATIONS\n"+tab+"/*checks the access from the request req under the context con*/\n"
    bmachine += tab+"access <-- CheckAccess(req, con)=\n"+tab+"PRE\n"+tab+tab+"req : REQUEST_T & con : CONTEXT\n"+tab + \
                "THEN\n"+tab+tab+"access := bool(access_def(req,con))\n"+tab+"END;\n\n"
    bmachine += tab+"/*checks if there is a hidden document under the context con*/\n"
    bmachine += tab+"hidden<-- HiddenDocument(con)=\n"+tab+"PRE\n"
    bmachine += tab+tab+"con : CONTEXT\n"+tab+"THEN\n"
    bmachine += tab+tab+"hidden := bool(#(hdoc).(hdoc:(V_RES - dom(e_res)) & !req.((req:REQUEST_T & " \
                        "prj2(V_SUB,V_RES)(req)=hdoc) => not(access_def(req,con)))))\n"+tab+"END;\n\n"
    bmachine += tab+"/*returns the set of the hidden documents under the condition con*/\n"
    bmachine += tab+"hiddenSet <-- HiddenDataSet(con)=\n"+tab+"PRE\n"
    bmachine += tab+tab+"con : CONTEXT\n"+tab+"THEN\n"
    bmachine += tab+tab+"hiddenSet := { hdoc | hdoc : V_RES - dom(e_res) & !req.((req:REQUEST_T & " \
                        "prj2(V_SUB,V_RES)(req)=hdoc) => not(access_def(req,con)))}\n"+tab+"END;\n\n"
    bmachine += tab+"/*determines the contexts that grant a request*/\n"
    bmachine += tab+"granting <-- GrantingContexts(req)=\n"+tab+"PRE\n"
    bmachine += tab+tab+"req : REQUEST_T\n"+tab+"THEN\n"
    bmachine += tab+tab+"granting := { con | con : CONTEXT & access_def(req,con)}\n"+tab+"END;\n\n"
    bmachine += tab+"/*returns the set of the rules that are never used*/\n"
    bmachine += tab+"ineffectiveSet <-- IneffectiveRuleSet =\n"+tab+"BEGIN\n"
    bmachine += tab+tab+"ineffectiveSet := {ru | ru : RULE_T & not(#(req,con).(req:REQUEST_T & ru : conRule(con) &" \
                        "ru : pseudoSink(req,con) &	(pseudoSink(req,con) - {ru} = {} or((rules(ru))'mo = pro & " \
                        "!ru2.(ru2:(pseudoSink(req,con)-{ru}) => (rules(ru2))'mo = per)))))}\n"+tab+"END\n"
    bmachine += "END"

    directory = "C:"+os.sep+"Users"+os.sep+"dead1401"+os.sep+"PycharmProjects"+os.sep+"SGAC"+os.sep+\
                "Tests"+os.sep+"varying_the_number_of_rules"+os.sep+"30rules"

    onlyfiles = next(os.walk(directory))[2]
    bMachineName = "SGAC_B_"+str((len(onlyfiles)) // 8 +1)+".mch"
    f = open(directory + os.sep + bMachineName, "w+")
    f.write(bmachine)
    f.close()

    timeB = time.time()
    p = subprocess.Popen("C:"+os.sep+"ProB"+os.sep+"probcli.exe -animate 10" + directory + os.sep + bMachineName,
                         stdout=subprocess.PIPE,
                         stderr=subprocess.PIPE,
                         shell=True)
    output, errors = p.communicate()
    n = ""
    timeB = time.time() - timeB
    print(timeB)
    if True:
        if p.returncode==0:
            print("ProB executed successfully ("+n+")")
            print(output)

        else:
            print("ProB - error reported in "+n+" and the return code is "+str(p.returncode))
            print(errors)

    timeP = time.time()

    f = open(directory + os.sep + "SGAC_random_"+str((len(onlyfiles)) // 8+1)+".txt", "w+")
    f.write(str(Python_Sub_Graph)), f.write("\n")
    f.write(str(Python_Res_Graph)), f.write("\n")
    f.write(str(Python_Context)), f.write("\n")
    f.write(str(Python_Rules)), f.write("\n")
    f.write(str(Python_Sub_Closure_Graph)), f.write("\n")
    f.write(str(Python_Res_Closure_Graph))
    f.close()

    SGAC.SGAC_random("SGAC_Z3_"+str(len(onlyfiles) // 8 + 1), directory, (len(onlyfiles)) // 8+1)

    timeP = time.time() - timeP
    print(timeP)

    resourceSink = []
    subjectSink = []
    requests = []
    for key in Python_Sub_Graph.keys():
        if not Python_Sub_Graph[key]:
            subjectSink.append(key)

    for key in Python_Res_Graph.keys():
        if not Python_Res_Graph[key]:
            resourceSink.append(key)

    for sub in subjectSink:
        for res in resourceSink:
            requests.append([sub, res])

    directory = "C:"+os.sep+"Users"+os.sep+"dead1401"+os.sep+"PycharmProjects"+os.sep+"SGAC"+os.sep+"Tests"+os.sep+\
                "varying_the_number_of_rules"+os.sep+"30rules"

    os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1))

    counter = 0
    for request in requests:
        os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1)+os.sep+"graph"+str(counter))
        alloyGraph = ""
        alloyGraph += "module filepath/param/graph/property/req\nopen filepath/alloy"+str(len(onlyfiles) // 8 + 1)+\
                      "/sgac_core\n//**********************\n" \
                      "//***Graph signatures***\n//**********************\n"
        alloyGraph += "one sig " + V_SUB[:len(V_SUB) - 2:] + " extends Subject{}{}\n"
        alloyGraph += "fact{\nsubSucc=" + graph_sub[:len(graph_sub) - 2:].replace(",", "+\n" + tab + tab).replace("|->",
                                                                                                                  "->") + \
                      "}\n"
        alloyGraph += "one sig " + V_RES[:len(V_RES) - 2:] + " extends Resource{}{}\n"
        alloyGraph += "fact{\nresSucc=" + graph_res[:len(graph_res) - 2:].replace(",", "+\n" + tab + tab).replace("|->",
                                                                                                                  "->") +\
                      "}\n"
        alloyGraph += "\n//*************************\n//***Contexts signatures***\n//*************************\n"
        alloyGraph += "one sig " + CONTEXT[:len(CONTEXT) - 2:] + " extends Context{}{}\n\n"
        alloyGraph += "//************************\n//***Request signatures***\n//************************\n"
        alloyGraph += "one sig req"+str(counter)+" extends Request{}{\n"
        alloyGraph += "sub="+request[0]+"\nres="+request[1]+"\n}\n"
        alloyGraph += "//**********************\n//***      Rules     ***\n//**********************\n"

        for eachRule in rulesInAlloy:
            for element in eachRule: \
                    alloyGraph += element + "\n"
            alloyGraph += "}\n"
        alloyGraph += "//**************************\n//***Auxiliary Predicates***\n//**************************\n" \
                      "pred access_condition[req:Request,con:Context]{\n" \
                      "" + tab + "(no  r:applicableRules[req] |  (evalRuleCond[r,con] and r.m=prohibition and\n" \
                                 "" + tab + tab + "all rule: r.^(req.ruleSucc) | not evalRuleCond[rule,con])	)\n" \
                                                  "" + tab + "and some { r:applicableRules[req] " \
                                                             "|evalRuleCond[r,con]}\n}\n\n"

        alloyAccess = alloyGraph + "//*********************\n//***Access property***\n//*********************\n"

        for con in Python_Context:
            alloyAccess += "run accessReq" + str(counter) + "_" +\
                           con+"{access_condition[req"+str(counter)+","+con+"]} for 4\n"
        os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "access")
        f = open(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "access" + os.sep + "req" + str(counter) + ".als",
                 "w+")
        f.write(alloyAccess)
        f.close()

        alloyContexts = alloyGraph+"//***************************\n//***Determination of the ***\n" \
                                  "//***conditions (contexts)***\n" \
                        "//***************************\n\none sig GrantingContext{\n" \
                        ""+tab+tab+"acc: set Context\n}{}\n\n"\
                        "pred grantingContextDet[req:Request]{\n"\
                        ""+tab+tab+"all c: Context | access_condition[req,c] <=> c in GrantingContext.acc\n}\n"\
                        "//*** determination command ***\n" \
                        "run grantingContextDetermination{grantingContextDet[req"+\
                        str(counter)+"]} for 4 but 1 GrantingContext"
        os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "contexts")
        f = open(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "contexts" + os.sep + "req" + str(counter)+".als",
                 "w+")
        f.write(alloyContexts)
        f.close()

        alloyHidden = alloyGraph+"//**************************\n//***Hidden data property***\n" \
                      "//**************************\n\nfun documents[res:Resource]: set Resource{\n" \
                      ""+tab+"{ rt : Resource | rt in res.^resSucc and no rt.resSucc}\n}\n\nfun documentsG[]: set Resource{" \
                      ""+tab+"{ rt : Resource | no rt.resSucc}}\n\nfun persons[s:Subject]: set Subject{\n" \
                      ""+tab+"{ sub: Subject | sub in s.^subSucc and no sub.subSucc}\n}\n\nfun personsG[]: set Subject{\n" \
                      ""+tab+"{ sub: Subject | no sub.subSucc}\n}\n\npred HiddenDocument[reso:Resource,c:Context]{\n" \
                      ""+tab+"no req: Request | (req.res = reso and\n"+tab+"access_condition[req,c])\n}\n\n" \
                      ""+tab+"pred HiddenData[c:Context] {\n"+tab+"some reso: documentsG[] | HiddenDocument[reso,c]\n}\n" \
                      "//***Hidden Data Existence and determination***\n"
        for con in Python_Context:
            alloyHidden += "check HiddenDocument_"+\
                           request[1]+"_"+con+"{ HiddenDocument["+request[1]+","+con+"]} for 4\n"
        os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "hidden")
        f = open(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "hidden" + os.sep + "req" + str(counter) + ".als",
                 "w+")
        f.write(alloyHidden)
        f.close()

        alloyIneffective = alloyGraph + "//***************************\n//***Determination of the ***\n" \
                           "//***  ineffective rules  ***\n//***************************\n\n" \
                           "fun pseudosinkRule[req: Request, cx:Context] : set Rule{\n" \
                           ""+tab+"{r : applicableRules[req] |\n" \
                           ""+tab+tab+"evalRuleCond[r,cx] and no ru : applicableRules[req] |\n" \
                           ""+tab+tab+tab+"ru in r.^(req.ruleSucc) and evalRuleCond[ru,cx]}\n" \
                           ""+tab+"}\n\n" \
                           "pred ineffectiveRule[r:Rule]{\n" \
                           ""+tab+"no rq : Request | (\n" \
                           ""+tab+tab+"r in applicableRules[rq] and some cr : r.c | (\n" \
                           ""+tab+tab+tab+"r in pseudosinkRule[rq,cr]\n" \
                           ""+tab+tab+tab+"and\n" \
                           ""+tab+tab+tab+"(no ru : pseudosinkRule[rq,cr]-r |\n" \
                           ""+tab+tab+tab+tab+"r.m = ru.m)\n" \
                           ""+tab+tab+tab+"and\n" \
                           ""+tab+tab+tab+"(r.m = permission implies no (pseudosinkRule[rq,cr]-r))\n" \
                           ""+tab+tab+")\n" \
                           ""+tab+")\n}\n" \
                           "//*** determination of unusable rules command ***\n\n"
        for key in Python_Rules.keys():
            alloyIneffective += "check ineffectiveRule"+key+"{ ineffectiveRule["+key+"]}for 4\n\n"
        os.mkdir(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "ineffective")
        f = open(directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "ineffective" + os.sep + "req" + str(counter) +
                 ".als", "w+")
        f.write(alloyIneffective)
        f.close()

        counter += + 1
    alloyCore = "open util/relation\nmodule sgac_core\n////////////////////////////////////////////////////////\n" \
                "// 						SGAC 											 //\n" \
                "////////////////////////////////////////////////////////\n" \
                "sig Subject {\n" \
                "	subSucc : set Subject                  //  subject successors; covering pair (s1,s2) in subSucc" \
                " <=> s1 son of s2, then, s1 plus specific than s2; s2 father of s1\n}\nfact { acyclic[subSucc,Subject] }\n" \
                "\nsig Resource {\n	resSucc : set Resource                 // resource successors; covering pair\n}\n" \
                "fact { acyclic[resSucc,Resource] }\n\nabstract sig Modality {}\n" \
                "one sig prohibition, permission extends  Modality {}\nsig Context {}\nsig Rule {\n" \
                "	p : one Int,         	 // priority of the rule\n	s : one Subject,\n	t : one Resource,\n" \
                "	m : one Modality,\n	c : set Context 			// truth table of the condition\n}{\n" \
                "  p >= 0\n}\n\nsig Request{\nsub: one Subject,\nres: one Resource,\nruleSucc:  Rule -> set Rule\n}{\n" \
                "no sub.subSucc\nno res.resSucc\nRule.ruleSucc in applicableRules[this]\n" \
                "ruleSucc.Rule in applicableRules[this]\n}\n\nfact {\n// definition of ruleSucc r1 -< r2\n" \
                "//    for Rule\n	all req: Request |\n		all r1,r2 : applicableRules[req] |\n" \
                "			r1 in req.ruleSucc.r2 // r1 -> r2 r2 plus prioritaire donc\n			<=>\n" \
                "           isPrecededBy[r1,r2]\n //        and\n" \
                " //          not (some r3 : applicableRules[req] | isPrecededBy[r1,r3] and isPrecededBy[r3,r2] )\n}\n\n" \
                "////////////////////////////////////////////////////////\n" \
                "// 							FORMULAS									 //\n" \
                "////////////////////////////////////////////////////////\n\n\npred evalRuleCond[r:Rule,con:Context]{\n" \
                "con in r.c\n}\n\n////////////////////////////////////////////////////////\n" \
                "// 						AUXILIARY FUNCTIONS					 //\n" \
                "////////////////////////////////////////////////////////\n\n" \
                "pred lessSpecific[r1,r2: Rule]{ // lessSpecific[r1,r2]=TRUE => " \
                "r2 has precedence over r1 (subject wise or priority)\n" \
                "	(r2.p < r1.p)\n	or\n	(\n		r2.p = r1.p\n			and\n		r2.s in (r1.s).^subSucc\n	)\n}\n" \
                "/*\npred applicableRule[r:Rule]{\nRequest.sub in (r.s).*subSucc\nand Request.res in (r.t).*resSucc\n}\n" \
                "*/\n\nfun applicableRules[req:Request]: set Rule{\n" \
                "{r: Rule | req.sub in (r.s).*subSucc and req.res in (r.t).*resSucc}\n}\n\n\npred maximal[r:Rule]{\n" \
                "no r1 : Rule | lessSpecific[r,r1]\n}\n\npred isPrecededBy[r1,r2:Rule]{\n	(\n 		lessSpecific[r1,r2]" \
                "\n		or\n		(\n			not lessSpecific[r2,r1]\n			and\n			maximal[r1]\n" \
                "			and\n			maximal[r2]\n			and\n			r2.m = prohibition\n" \
                "			and\n			r1.m != r2.m\n		)\n	)\n}\n\npred sinkRule[r:Rule,req:Request]{\n" \
                "no r.(req.ruleSucc)\n}\n\n\nrun{} for exactly 4 Rule, 4 Subject, 4 Resource, 4 Context, 2 Request\n\n" \
                "assert acyclicRule{ all r: Request | acyclic[r.ruleSucc,applicableRules[r]]}\n" \
                "check acyclicRule for  7 but exactly 5 Rule"

    f = open(directory +os.sep+"alloy"+str(len(onlyfiles) // 8 + 1)+ os.sep + "sgac_core.als", "w")
    f.write(alloyCore)
    f.close()

    counter = counter - 1
    timeAComulative = 0
    for request in requests:
        timeAInit = time.time()
        p = subprocess.Popen("cd C:"+os.sep+"Users"+os.sep+"dead1401"+os.sep+"Documents"+os.sep+"alloy"+os.sep+
                             "tool & java -Xss4m -cp \".;alloy4.2_2015-02-22.jar\" ExampleUsingTheCompiler "+
                             directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "access" + os.sep + "req" +
                             str(counter) + ".als "+
                             directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "contexts" + os.sep + "req" +
                             str(counter) + ".als "+
                             directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "hidden" + os.sep + "req" +
                             str(counter) + ".als "+
                             directory+os.sep+"alloy"+str(len(onlyfiles) // 8 + 1) + os.sep + "graph" + str(counter) + os.sep + "ineffective" + os.sep + "req" +
                             str(counter) + ".als ",
                             stdout=subprocess.PIPE,
                             stderr=subprocess.PIPE,
                             shell=True)
        output, errors = p.communicate()
        n = ""
        timeAComulative += time.time() - timeAInit
        if True:
            if p.returncode==0:
                print("Alloy Analyzer executed successfully ("+n+")")
                print(output)
            else:
                print("Alloy Analyzer - error reported in "+n+" and the return code is "+str(p.returncode))
                print(errors)
        counter -= 1
    print(timeAComulative)

    directory = "C:"+os.sep+"Users"+os.sep+"dead1401"+os.sep+"PycharmProjects"+os.sep+"SGAC"+os.sep+"Tests"+os.sep+\
                "varying_the_number_of_rules"

    if os.path.exists(directory+os.sep+'times30rules.xlsx'):
        wb = load_workbook(directory+os.sep+'times30rules.xlsx')
        ws = wb.worksheets[0]
        row_count = ws.max_row
        d = ws.cell(row=row_count+1, column=1, value=timeB)
        d = ws.cell(row=row_count+1, column=2, value=timeP)
        d = ws.cell(row=row_count+1, column=3, value=timeAComulative)
        wb.save(directory+os.sep+'times30rules.xlsx')
    else:
        wb = Workbook()
        ws = wb.active
        ws1 = wb.create_sheet("Times")
        ws.title = "SGAC Execution Time"
        ws1 = wb["SGAC Execution Time"]
        d = ws.cell(row=1, column=1, value="ProB")
        d = ws.cell(row=1, column=2, value="Z3")
        d = ws.cell(row=1, column=3, value="Alloy")
        d = ws.cell(row=2, column=1, value=timeB)
        d = ws.cell(row=2, column=2, value=timeP)
        d = ws.cell(row=2, column=3, value=timeAComulative)
        wb.save(directory+os.sep+'times30rules.xlsx')
